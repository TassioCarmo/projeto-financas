"""Leitor para planilhas multi-aba no formato Controle Cartão Familiar."""

import re
from io import BytesIO
from typing import BinaryIO

import openpyxl
import pandas as pd

from app.servicos.perfis_importacao import normalizar_nome_coluna

MESES = (
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
)

MESES_NORMALIZADOS = {normalizar_nome_coluna(mes): mes for mes in MESES}

COLUNAS_ESPERADAS = frozenset(
    {
        normalizar_nome_coluna("Descrição da Compra"),
        normalizar_nome_coluna("Valor (R$)"),
    }
)

_PREFIXOS_RESUMO = ("RESUMO", "TOTAL", "Total", "Qtd.")

_ABA_MES_COL = "_aba_mes"
_ABA_ANO_COL = "_aba_ano"


def _cabecalho_cartao_familiar(linha: tuple) -> bool:
    nomes = {normalizar_nome_coluna(c) for c in linha if c is not None}
    return COLUNAS_ESPERADAS.issubset(nomes)


def detectar_formato_cartao_familiar(arquivo: BinaryIO) -> bool:
    """Verifica se o XLSX segue o layout Controle Cartão Familiar."""
    conteudo = arquivo.read()
    arquivo.seek(0)
    wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    try:
        for nome in wb.sheetnames:
            mes_norm = normalizar_nome_coluna(nome)
            if mes_norm not in MESES_NORMALIZADOS:
                continue
            ws = wb[nome]
            linha_cabecalho = next(
                ws.iter_rows(min_row=2, max_row=2, values_only=True),
                None,
            )
            if linha_cabecalho and _cabecalho_cartao_familiar(linha_cabecalho):
                return True
        return False
    finally:
        wb.close()


def _extrair_ano_titulo(valor) -> int | None:
    if valor is None:
        return None
    match = re.search(r"(20\d{2})", str(valor))
    return int(match.group(1)) if match else None


def _linha_resumo(descricao) -> bool:
    if descricao is None or (isinstance(descricao, float) and pd.isna(descricao)):
        return False
    texto = str(descricao).strip()
    if not texto:
        return False
    if "⚠️" in texto or "⚠" in texto:
        return True
    return any(texto.startswith(prefixo) for prefixo in _PREFIXOS_RESUMO)


def _filtrar_linhas(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    def _linha_deve_ser_removida(row: pd.Series) -> bool:
        for col, valor in row.items():
            if col in (_ABA_MES_COL, _ABA_ANO_COL):
                continue
            if _linha_resumo(valor):
                return True
        return all(
            valor is None or (isinstance(valor, float) and pd.isna(valor)) or str(valor).strip() == ""
            for col, valor in row.items()
            if col not in (_ABA_MES_COL, _ABA_ANO_COL)
        )

    return df[~df.apply(_linha_deve_ser_removida, axis=1)].reset_index(drop=True)


def ler_planilha_cartao_familiar(arquivo: BinaryIO) -> pd.DataFrame:
    """Lê e concatena as abas mensais de um XLSX Controle Cartão Familiar."""
    conteudo = arquivo.read()
    arquivo.seek(0)

    wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    abas_mensais: list[str] = []
    anos_por_aba: dict[str, int | None] = {}
    try:
        for nome in wb.sheetnames:
            mes_norm = normalizar_nome_coluna(nome)
            if mes_norm not in MESES_NORMALIZADOS:
                continue
            ws = wb[nome]
            linha_cabecalho = next(
                ws.iter_rows(min_row=2, max_row=2, values_only=True),
                None,
            )
            if not linha_cabecalho or not _cabecalho_cartao_familiar(linha_cabecalho):
                continue
            abas_mensais.append(nome)
            titulo = ws.cell(row=1, column=1).value
            anos_por_aba[nome] = _extrair_ano_titulo(titulo)
    finally:
        wb.close()

    if not abas_mensais:
        return pd.DataFrame()

    partes: list[pd.DataFrame] = []
    for nome in abas_mensais:
        df_aba = pd.read_excel(
            BytesIO(conteudo),
            sheet_name=nome,
            header=1,
            engine="openpyxl",
        )
        df_aba[_ABA_MES_COL] = MESES_NORMALIZADOS[normalizar_nome_coluna(nome)]
        df_aba[_ABA_ANO_COL] = anos_por_aba.get(nome)
        partes.append(_filtrar_linhas(df_aba))

    if not partes:
        return pd.DataFrame()

    return pd.concat(partes, ignore_index=True)
